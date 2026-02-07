import openpyxl
import os

def format_template_dates(file_path='wzór.xlsx'):
    if not os.path.exists(file_path):
        print(f"Błąd: Plik {file_path} nie istnieje.")
        return

    print(f"Wczytywanie {file_path}...")
    wb = openpyxl.load_workbook(file_path)
    ws = wb.active

    # Kolumna E ("Data faktury") to 5. kolumna
    # Kolumna P ("Kwota rachunku") to 16. kolumna
    # Kolumna R ("Kwota do wypłaty") to 18. kolumna
    cols_to_format = {
        5: 'YYYY-MM-DD',
        16: '0.00',
        18: '0.00'
    }
    
    print(f"Ustawianie formatowania dla wybranych kolumn...")
    
    # Przetwarzamy wszystkie wiersze z danymi (od 2 w górę)
    updated_count = 0
    for row in range(2, ws.max_row + 1):
        for col_idx, fmt in cols_to_format.items():
            cell = ws.cell(row=row, column=col_idx)
            cell.number_format = fmt
        updated_count += 1

    wb.save(file_path)
    print(f"Sukces! Zaktualizowano formatowanie w {updated_count} wierszach.")

if __name__ == "__main__":
    format_template_dates()
