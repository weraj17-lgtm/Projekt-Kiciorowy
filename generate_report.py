import openpyxl
from openpyxl.styles import Font, PatternFill
from datetime import datetime
import sys
import os
import unicodedata
import math

def get_roman(n):
    """Zamienia liczbę na cyfrę rzymską (do części raportu)."""
    romans = {
        1: 'I', 2: 'II', 3: 'III', 4: 'IV', 5: 'V', 
        6: 'VI', 7: 'VII', 8: 'VIII', 9: 'IX', 10: 'X',
        11: 'XI', 12: 'XII', 13: 'XIII', 14: 'XIV', 15: 'XV',
        16: 'XVI', 17: 'XVII', 18: 'XVIII', 19: 'XIX', 20: 'XX'
    }
    return romans.get(n, str(n))

def clean_non_polish_chars(text):
    """
    Zamienia znaki niepolskie na ich odpowiedniki, zachowując polskie znaki.
    """
    if not isinstance(text, str):
        return text
    
    # Znaki, które CHCEMY zachować (polskie + standardowe)
    polish_chars = "ąćęłńóśźżĄĆĘŁŃÓŚŹŻ"
    allowed_chars = set(polish_chars + "abcdefghijklmnopqrstuvwxyzABCDEFGHIJKLMNOPQRSTUVWXYZ0123456789 ,.-_/()[]:;\"'?!@#$%^&*+=|<>")
    
    result = []
    for char in text:
        if char in allowed_chars:
            result.append(char)
        else:
            # Próbujemy normalizacji dla dekompozycji (np. ä -> a + kropki)
            unicodedata.normalize('NFKD', char)
            normalized = unicodedata.normalize('NFKD', char)
            base_char = normalized[0]
            if base_char in allowed_chars:
                result.append(base_char)
            else:
                pass
                
    return "".join(result)

def process_worksheet(ws, report_date, headers):
    """Wykonuje wszystkie transformacje na arkuszu."""
    try:
        col_data_idx = headers.index('Data raportu') + 1
        col_lp_idx = headers.index('L.p.') + 1
        col_nk_idx = headers.index('NK') + 1 if 'NK' in headers else None
        col_faktura_idx = headers.index('Nr faktury') + 1 if 'Nr faktury' in headers else None
        col_r_idx = headers.index('Kwota do wypłaty') + 1 if 'Kwota do wypłaty' in headers else 18
        col_h_idx = headers.index('Stanowisko Kosztowe') + 1 if 'Stanowisko Kosztowe' in headers else 8
        col_t_idx = 20 # Kolumna T
    except ValueError as e:
        print(f"Błąd: Nie znaleziono wymaganych kolumn: {e}")
        raise

    # Definicja stylów
    red_font = Font(color="FF0000") # Czerwony
    green_fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid") # Jasnozielony (standard Excel)

    # Przygotuj datę
    new_date_val = None
    if report_date:
        try:
            new_date_val = datetime.strptime(report_date, "%d.%m.%Y")
        except:
            new_date_val = report_date

    # Przetwarzaj wiersze
    row_count = 0
    nk_updated_count = 0
    faktura_updated_count = 0
    t_updated_count = 0
    high_value_count = 0
    green_row_count = 0
    
    max_r = ws.max_row
    
    for row in range(2, max_r + 1):
        # 1. Zaktualizuj datę raportu
        if report_date:
            cell_date = ws.cell(row=row, column=col_data_idx)
            cell_date.value = new_date_val
            cell_date.number_format = 'DD.MM.YYYY'
            
        # 2. Wyczyść kolumnę L.p.
        ws.cell(row=row, column=col_lp_idx).value = None

        # 3. Przetwarzaj kolumnę NK
        if col_nk_idx:
            cell_nk = ws.cell(row=row, column=col_nk_idx)
            try:
                val_raw = cell_nk.value
                val = float(val_raw) if val_raw is not None else 0
                if val > 700:
                    cell_nk.value = "700"
                    cell_nk.number_format = '@'
                    nk_updated_count += 1
            except (ValueError, TypeError):
                pass
        
        # 4. Przetwarzaj kolumnę Nr faktury
        invoice_highlights_row = False
        if col_faktura_idx:
            cell_faktura = ws.cell(row=row, column=col_faktura_idx)
            if cell_faktura.value is not None:
                orig_val = str(cell_faktura.value)
                new_val = orig_val.replace('\\', '/').replace('_', '/').replace(';', '/')
                if new_val != orig_val:
                    cell_faktura.value = new_val
                    faktura_updated_count += 1
                
                # Sprawdź czy kończy się na wzorce (strip usuwa spacje)
                check_val = new_val.strip()
                if any(check_val.endswith(s) for s in ['00/25', '00/26', '00/2025', '00/2026']):
                    invoice_highlights_row = True

        # 5. Przetwarzaj kolumnę T
        cell_t = ws.cell(row=row, column=col_t_idx)
        if cell_t.value is not None:
            old_val = str(cell_t.value)
            new_val = clean_non_polish_chars(old_val)
            if new_val != old_val:
                cell_t.value = new_val
                t_updated_count += 1

        # 6. Wyróżnianian kwot >= 500 w kolumnie R
        cell_r = ws.cell(row=row, column=col_r_idx)
        try:
            val_r = float(cell_r.value) if cell_r.value is not None else 0
            if val_r >= 500:
                cell_r.font = red_font
                high_value_count += 1
        except (ValueError, TypeError):
            pass

        # 7. Format tekstowy dla kolumny H (Stanowisko Kosztowe)
        if col_h_idx:
            cell_h = ws.cell(row=row, column=col_h_idx)
            cell_h.number_format = '@'

        # 8. Jeśli wiersz ma być zielony (faktura)
        if invoice_highlights_row:
            green_row_count += 1
            for col in range(1, ws.max_column + 1):
                ws.cell(row=row, column=col).fill = green_fill

        row_count += 1

    # Usuń kolumny V (22) i W (23) przed agregacją
    ws.delete_cols(23)
    ws.delete_cols(22)

    # Wskazówki kolumn dla agregacji (po usunięciu V i W)
    col_beneficjent_idx = 19 # S
    col_kod_idx = 14 # N
    
    # --- AGREGACJA ---
    aggregation = {}
    
    # Przetwarzamy wiersze danych (od 2 do row_count+1)
    for r in range(2, ws.max_row + 1):
        def get_val(c_idx):
            v = ws.cell(row=r, column=c_idx).value
            return str(v).strip() if v is not None else ""

        acc_num = get_val(col_r_idx or 18)
        beneficjent = get_val(col_beneficjent_idx)
        kod = get_val(col_kod_idx)
        
        if not acc_num and not beneficjent:
            continue
            
        key = (acc_num, beneficjent, kod)
        
        # Dane do agregacji
        try:
            kwota = float(ws.cell(row=r, column=col_r_idx).value or 0)
        except:
            kwota = 0
            
        fv = get_val(col_faktura_idx)
        
        if key not in aggregation:
            aggregation[key] = {
                'sum_kwota': kwota,
                'faktury': [fv] if fv else [],
                'patient_name': f"{get_val(8)} {get_val(9)}".strip()
            }
        else:
            aggregation[key]['sum_kwota'] += kwota
            if fv and fv not in aggregation[key]['faktury']:
                aggregation[key]['faktury'].append(fv)

    # Czyszczenie arkusza i wpisywanie wyników (zastępujemy dane)
    ws.delete_rows(2, ws.max_row)
    
    # Nowe nagłówki dla arkusza wynikowego (zgodnie z makrem)
    new_headers = ["Suma Kwoty", "Beneficjent", "Nr Konta", "konto bankowe z którego idą płatności", "Opis nr fv", "data realizacji"]
        
    for i, h in enumerate(new_headers, 1):
        ws.cell(row=1, column=i).value = h
        ws.cell(row=1, column=i).font = openpyxl.styles.Font(bold=True)

    for i, (key, data) in enumerate(aggregation.items(), 2):
        ws.cell(row=i, column=1).value = data['sum_kwota']
        ws.cell(row=i, column=2).value = data['patient_name']
        ws.cell(row=i, column=3).value = key[0] # Konto
        ws.cell(row=i, column=3).number_format = '@'
        ws.cell(row=i, column=4).value = "" # Konto bankowe z którego idą płatności ma być puste
        ws.cell(row=i, column=5).value = ", ".join(data['faktury'])
        ws.cell(row=i, column=5).number_format = '@'
        ws.cell(row=i, column=6).value = "" # Data realizacji ma być pusta

    # Dodatkowo usuń po nazwie (jeśli jeszcze zostały lub są w innych miejscach)
    current_headers = [str(cell.value).strip() if cell.value else "" for cell in ws[1]]
    cols_to_delete = ['Skompletował', 'Skompletował - IDX']
    for col_name in reversed(cols_to_delete):
        if col_name in current_headers:
            idx = current_headers.index(col_name) + 1
            ws.delete_cols(idx)
            print(f"  - Usunięto kolumnę: '{col_name}'")

    return {
        'rows': len(aggregation),
        'nk': nk_updated_count,
        'faktury': faktura_updated_count,
        't_chars': t_updated_count,
        'high_values': high_value_count,
        'green_rows': green_row_count
    }

def generate_report(input_file='wzór.xlsx', output_dir=None, report_date=None):
    """
    Generuje raport refundacji, dzieląc go na części jeśli przekracza 2000 wierszy.
    """
    print(f"Wczytywanie danych z: {input_file}")
    try:
        wb_template = openpyxl.load_workbook(input_file)
    except Exception as e:
        print(f"Błąd podczas wczytywania pliku: {e}")
        raise
        
    ws_template = wb_template.active
    headers = [str(cell.value).strip() if cell.value else "" for cell in ws_template[1]]
    
    # Wykonaj główne przetwarzanie na tym skoroszycie
    stats = process_worksheet(ws_template, report_date, headers)
    
    total_data_rows = ws_template.max_row - 1
    rows_per_part = 2000
    
    if total_data_rows <= 0:
        print("Brak danych do przetworzenia.")
        return None

    num_parts = math.ceil(total_data_rows / rows_per_part)
    
    if output_dir is None:
        output_dir = os.path.dirname(os.path.abspath(input_file))

    base_output_name = f"Raport Refundacje {report_date}" if report_date else "Raport Refundacje"
    
    results = []
    
    if num_parts <= 1:
        output_path = os.path.join(output_dir, f"{base_output_name}.xlsx")
        wb_template.save(output_path)
        print(f"\n[OK] Raport wygenerowany: {base_output_name}.xlsx ({total_data_rows} wierszy)")
        results.append(output_path)
    else:
        print(f"\nLiczba wierszy ({total_data_rows}) przekracza {rows_per_part}. Dzielenie na {num_parts} części...")
        
        temp_processed_path = os.path.join(output_dir, "temp_processed.xlsx")
        wb_template.save(temp_processed_path)
        
        for p in range(1, num_parts + 1):
            roman_part = get_roman(p)
            part_filename = f"{base_output_name} cz. {roman_part}.xlsx"
            part_path = os.path.join(output_dir, part_filename)
            
            # Wczytaj pełny skoroszyt ze wszystkimi arkuszami
            wb_part = openpyxl.load_workbook(temp_processed_path)
            
            # Pobierz aktywny arkusz (główny z danymi)
            ws_part = wb_part.active
            active_sheet_name = ws_part.title
            
            start_row = (p - 1) * rows_per_part + 2
            end_row = min(p * rows_per_part + 1, total_data_rows + 1)
            
            # Usuń wiersze poza zakresem tylko w aktywnym arkuszu
            if end_row < ws_part.max_row:
                ws_part.delete_rows(end_row + 1, ws_part.max_row - end_row)
            
            if start_row > 2:
                ws_part.delete_rows(2, start_row - 2)
            
            # Wszystkie pozostałe arkusze są automatycznie zachowane
            sheet_names = [s for s in wb_part.sheetnames if s != active_sheet_name]
            if sheet_names:
                print(f"    (zachowano arkusze: {', '.join(sheet_names)})")
            
            wb_part.save(part_path)
            print(f"  - Zapisano: {part_filename} (wiersze {start_row-1}-{end_row-1})")
            results.append(part_path)
            
        if os.path.exists(temp_processed_path):
            os.remove(temp_processed_path)
            
    print(f"\nStatystyki ogólne:")
    print(f"  - Przetworzono wierszy: {stats['rows']}")
    print(f"  - Poprawiono NK: {stats['nk']}")
    print(f"  - Poprawiono Nr faktur: {stats['faktury']}")
    print(f"  - Oczyszczono znaki w kolumnie T: {stats['t_chars']}")
    print(f"  - Wyróżniono kwot > 500: {stats['high_values']}")
    print(f"  - Wyróżniono wierszy (faktury 00/25-26): {stats['green_rows']}")
    
    return results

def main():
    """Główna funkcja skryptu."""
    input_file = 'wzór.xlsx'
    report_date = None
    
    if len(sys.argv) > 1:
        if sys.argv[1].endswith('.xlsx'):
            input_file = sys.argv[1]
            if len(sys.argv) > 2:
                report_date = sys.argv[2]
        else:
            report_date = sys.argv[1]
    
    if report_date is None:
        print("\nPodaj datę do nazwy pliku (np. 31.01.2026) lub naciśnij Enter:")
        try:
            report_date = input("> ").strip()
        except EOFError:
            report_date = ""
        if not report_date:
            report_date = None
    
    if not os.path.exists(input_file):
        script_dir = os.path.dirname(os.path.abspath(__file__))
        alt_path = os.path.join(script_dir, input_file)
        if os.path.exists(alt_path):
            input_file = alt_path
        else:
            print(f"Błąd: Plik '{input_file}' nie istnieje!")
            sys.exit(1)
    
    try:
        generate_report(input_file, report_date=report_date)
    except Exception as e:
        print(f"Wyjątek: {e}")
        import traceback
        traceback.print_exc()
        sys.exit(1)

if __name__ == "__main__":
    main()
