"""
Skrypt do uzupełniania pliku wzór.xlsx przykładowymi danymi.
Dodaje 2000 wierszy na wzór dwóch pierwszych wierszy danych.
"""

import openpyxl
import random
import os

def fill_template(file_path='wzór.xlsx', num_rows=2000):
    if not os.path.exists(file_path):
        print(f"Błąd: Plik {file_path} nie istnieje.")
        return

    print(f"Wczytywanie {file_path}...")
    wb = openpyxl.load_workbook(file_path)
    ws = wb.active

    # Pobierz pierwsze dwa wiersze danych (wiersz 2 i 3)
    # ws[1] to nagłówki, ws[2] i ws[3] to dane
    sample_rows = []
    for row_idx in [2, 3]:
        row_data = [cell.value for cell in ws[row_idx]]
        sample_rows.append(row_data)

    print(f"Generowanie {num_rows} wierszy...")
    
    # Znajdź obecną maksymalną wartość L.p. (kolumna C, index 2)
    current_lp = 0
    for row in range(2, ws.max_row + 1):
        val = ws.cell(row=row, column=3).value
        if isinstance(val, int):
            current_lp = max(current_lp, val)

    # Dodaj nowe wiersze
    for i in range(1, num_rows + 1):
        # Losowo wybierz jeden z wzorcowych wierszy
        new_row = list(random.choice(sample_rows))
        
        # Zaktualizuj L.p.
        current_lp += 1
        new_row[2] = current_lp # Kolumna L.p.
        
        # Opcjonalnie: zróżnicuj numery faktur, żeby nie były identyczne
        if new_row[3]: # Nr faktury
            new_row[3] = f"{new_row[3]}_test_{i}"

        ws.append(new_row)

    print(f"Zapisywanie zmian do {file_path}...")
    wb.save(file_path)
    print(f"Sukces! Dodano {num_rows} wierszy. Łączna liczba wierszy danych: {ws.max_row - 1}")

if __name__ == "__main__":
    fill_template()
